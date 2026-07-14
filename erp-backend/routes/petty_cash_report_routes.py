# petty_cash_report.py
import logging
from flask import Blueprint, request, jsonify, send_file
from models import db, PettyCash, PettyCashLedger, User, Branch, PettyCashFundAllocation
from helpers import token_required, get_user_allowed_branches
from datetime import datetime
from sqlalchemy import extract, func
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

petty_cash_report_bp = Blueprint('petty_cash_report_bp', __name__)
logger = logging.getLogger(__name__)


# -----------------------------
# BRANCH WISE EXPENSES (View 1)
# -----------------------------
@petty_cash_report_bp.route('/branch-wise', methods=['GET'])
@token_required
def branch_wise_expenses(current_user):
    try:
        academic_year = request.headers.get("X-Academic-Year", "2024-2025")

        # Determine FY start/end year
        if '-' in academic_year:
            fy_start_year = int(academic_year.split('-')[0])
        else:
            fy_start_year = int(academic_year)

        # FY runs from April (month 4) of fy_start_year to March (month 3) of fy_start_year+1
        months = [
            ("Apr", fy_start_year, 4),
            ("May", fy_start_year, 5),
            ("Jun", fy_start_year, 6),
            ("Jul", fy_start_year, 7),
            ("Aug", fy_start_year, 8),
            ("Sep", fy_start_year, 9),
            ("Oct", fy_start_year, 10),
            ("Nov", fy_start_year, 11),
            ("Dec", fy_start_year, 12),
            ("Jan", fy_start_year + 1, 1),
            ("Feb", fy_start_year + 1, 2),
            ("Mar", fy_start_year + 1, 3),
        ]

        # Get all branches accessible to the user
        allowed = get_user_allowed_branches(current_user)
        if allowed['is_unlimited']:
            branches = Branch.query.filter_by(is_active=True).all()
        elif allowed['ids']:
            branches = Branch.query.filter(Branch.id.in_(allowed['ids']), Branch.is_active==True).all()
        else:
            branches = []

        result = []
        for branch in branches:
            row = {
                "branch_id": branch.id,
                "branch_name": branch.branch_name,
                "months": {},
                "total": 0.0
            }
            for label, yr, mn in months:
                total = db.session.query(func.coalesce(func.sum(PettyCash.amount), 0))\
                    .filter(
                        PettyCash.branch_id == branch.id,
                        PettyCash.academic_year == academic_year,
                        PettyCash.is_active == True,
                        PettyCash.approval_status == 'Approved',
                        extract('month', PettyCash.transaction_date) == mn,
                        extract('year', PettyCash.transaction_date) == yr
                    ).scalar()
                amt = float(total or 0)
                row["months"][label] = amt
                row["total"] += amt
            result.append(row)

        return jsonify(result), 200
    except Exception as e:
        logger.error(f"Error in branch wise expenses: {str(e)}")
        return jsonify({"message": str(e)}), 500


# -----------------------------
# LEDGER HEAD EXPENSES (View 2)
# -----------------------------
@petty_cash_report_bp.route('/ledger-head', methods=['GET'])
@token_required
def ledger_head_expenses(current_user):
    try:
        academic_year = request.headers.get("X-Academic-Year", "2024-2025")
        location = request.args.get("location")
        branch_id_param = request.args.get("branch_id")

        if '-' in academic_year:
            fy_start_year = int(academic_year.split('-')[0])
        else:
            fy_start_year = int(academic_year)
        
        months = [
            ("Apr", fy_start_year, 4),
            ("May", fy_start_year, 5),
            ("Jun", fy_start_year, 6),
            ("Jul", fy_start_year, 7),
            ("Aug", fy_start_year, 8),
            ("Sep", fy_start_year, 9),
            ("Oct", fy_start_year, 10),
            ("Nov", fy_start_year, 11),
            ("Dec", fy_start_year, 12),
            ("Jan", fy_start_year + 1, 1),
            ("Feb", fy_start_year + 1, 2),
            ("Mar", fy_start_year + 1, 3),
        ]

        query = PettyCash.query.filter(
            PettyCash.academic_year == academic_year,
            PettyCash.voucher_type.in_(['Payment', 'Payments']),
            PettyCash.is_active == True,
            PettyCash.approval_status == 'Approved'
        )

        allowed = get_user_allowed_branches(current_user)
        if branch_id_param and branch_id_param != 'All':
            query = query.filter(PettyCash.branch_id == int(branch_id_param))
        elif not allowed['is_unlimited']:
            if allowed['ids']:
                query = query.filter(PettyCash.branch_id.in_(allowed['ids']))
            else:
                query = query.filter(False)

        ledgers = PettyCashLedger.query.filter_by(is_active=True).all()

        result = []
        for ledger in ledgers:
            row = {
                "ledger_id": ledger.id,
                "ledger_head": ledger.ledger_name,
                "months": {},
                "total": 0.0
            }
            for label, yr, mn in months:
                total = query.filter(
                    PettyCash.ledger_id == ledger.id,
                    extract('month', PettyCash.transaction_date) == mn,
                    extract('year', PettyCash.transaction_date) == yr
                ).with_entities(func.coalesce(func.sum(PettyCash.amount), 0)).scalar()
                amt = float(total or 0)
                row["months"][label] = amt
                row["total"] += amt
            result.append(row)

        return jsonify(result), 200
    except Exception as e:
        logger.error(f"Error in ledger head expenses: {str(e)}")
        return jsonify({"message": str(e)}), 500


# -----------------------------
# BRANCH EXPENSE DETAILS (View 3)
# -----------------------------
@petty_cash_report_bp.route('/branch-details', methods=['GET'])
@token_required
def branch_expense_details(current_user):
    try:
        academic_year = request.headers.get("X-Academic-Year", "2024-2025")
        branch_id = request.args.get("branch_id")
        month = request.args.get("month")
        year = request.args.get("year")
        from_date = request.args.get("from_date")
        to_date = request.args.get("to_date")

        query = PettyCash.query.filter(
            PettyCash.academic_year == academic_year,
            PettyCash.voucher_type.in_(['Payment', 'Payments']),
            PettyCash.is_active == True,
            PettyCash.approval_status == 'Approved'
        )

        allowed = get_user_allowed_branches(current_user)
        if branch_id and branch_id != 'All':
            query = query.filter(PettyCash.branch_id == int(branch_id))
        elif not allowed['is_unlimited']:
            if allowed['ids']:
                query = query.filter(PettyCash.branch_id.in_(allowed['ids']))
            else:
                query = query.filter(False)

        if month and year:
            query = query.filter(
                extract('month', PettyCash.transaction_date) == int(month),
                extract('year', PettyCash.transaction_date) == int(year)
            )
        elif from_date and to_date:
            query = query.filter(
                PettyCash.transaction_date >= datetime.strptime(from_date, '%Y-%m-%d').date(),
                PettyCash.transaction_date <= datetime.strptime(to_date, '%Y-%m-%d').date()
            )

        transactions = query.order_by(PettyCash.transaction_date.asc()).all()

        result = []
        sno = 1
        for t in transactions:
            branch = Branch.query.get(t.branch_id)
            creator = User.query.get(t.created_by) if t.created_by else None
            ledger = PettyCashLedger.query.get(t.ledger_id) if t.ledger_id else None
            
            base_row = {
                "date": t.transaction_date.strftime('%d %b %Y'),
                "branch_name": branch.branch_name if branch else "",
                "voucher_no": t.voucher_name or f"C{t.id:05d}",
                "paying_account": f"{branch.branch_name if branch else ''} - Petty A/c ()" if branch else "",
                "voucher_type": t.voucher_type,
                "ledger_type": ledger.ledger_type if ledger else "",
                "ledger_head": ledger.ledger_name if ledger else "",
                "paid_to": t.paid_to or "",
                "description": t.description or "",
                "created_by": creator.username if creator else "",
                "approved_by": t.approved_by or "",
            }

            item_names = ", ".join([i.item_name for i in t.items]) if getattr(t, 'items', None) and len(t.items) > 0 else (t.voucher_name or "")
            row = dict(base_row)
            row["sno"] = sno
            row["ledger_name"] = item_names
            row["amount"] = float(t.amount)
            result.append(row)
            sno += 1

        return jsonify(result), 200
    except Exception as e:
        logger.error(f"Error in branch expense details: {str(e)}")
        return jsonify({"message": str(e)}), 500


# -----------------------------
# EXCEL EXPORT
# -----------------------------
@petty_cash_report_bp.route('/export-excel', methods=['GET'])
@token_required
def export_excel(current_user):
    try:
        academic_year = request.headers.get("X-Academic-Year", "2024-2025")
        branch_id = request.args.get("branch_id")
        month = request.args.get("month")
        year = request.args.get("year")
        from_date = request.args.get("from_date")
        to_date = request.args.get("to_date")

        query = PettyCash.query.filter(
            PettyCash.academic_year == academic_year,
            PettyCash.voucher_type.in_(['Payment', 'Payments']),
            PettyCash.is_active == True,
            PettyCash.approval_status == 'Approved'
        )

        allowed = get_user_allowed_branches(current_user)
        if branch_id and branch_id != 'All':
            query = query.filter(PettyCash.branch_id == int(branch_id))
        elif not allowed['is_unlimited']:
            if allowed['ids']:
                query = query.filter(PettyCash.branch_id.in_(allowed['ids']))
            else:
                query = query.filter(False)

        if month and year:
            query = query.filter(
                extract('month', PettyCash.transaction_date) == int(month),
                extract('year', PettyCash.transaction_date) == int(year)
            )
        elif from_date and to_date:
            query = query.filter(
                PettyCash.transaction_date >= datetime.strptime(from_date, '%Y-%m-%d').date(),
                PettyCash.transaction_date <= datetime.strptime(to_date, '%Y-%m-%d').date()
            )

        transactions = query.order_by(PettyCash.transaction_date.asc()).all()

        # Build workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Petty Cash Report"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right = Alignment(horizontal="right", vertical="center")

        headers = [
            "S.No", "Date", "Branch Name", "Voucher No", "Paying Account",
            "Voucher", "Ledger", "Ledger Head", "Items",
            "Paid To", "Description", "Created By", "Approved By", "Amount", "Total"
        ]

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        sno = 1
        row_idx = 2
        for t in transactions:
            branch = Branch.query.get(t.branch_id)
            creator = User.query.get(t.created_by) if t.created_by else None
            ledger = PettyCashLedger.query.get(t.ledger_id) if t.ledger_id else None
            
            items_to_process = t.items if getattr(t, 'items', None) and len(t.items) > 0 else [None]
            num_items = len(items_to_process)
            start_row = row_idx
            
            voucher_total = sum(float(item.amount) for item in t.items) if getattr(t, 'items', None) else float(t.amount)

            for i, item in enumerate(items_to_process):
                is_first = (i == 0)
                
                # Column 1 to 8
                ws.cell(row=row_idx, column=1, value=sno if is_first else "")
                ws.cell(row=row_idx, column=2, value=t.transaction_date.strftime('%d-%b-%Y') if is_first else "")
                ws.cell(row=row_idx, column=3, value=(branch.branch_name if branch else "") if is_first else "")
                ws.cell(row=row_idx, column=4, value=(t.voucher_name or f"C{t.id:05d}") if is_first else "")
                ws.cell(row=row_idx, column=5, value=(f"{branch.branch_name if branch else ''} - Petty A/c ()" if branch else "") if is_first else "")
                ws.cell(row=row_idx, column=6, value=t.voucher_type if is_first else "")
                ws.cell(row=row_idx, column=7, value=(ledger.ledger_type if ledger else "") if is_first else "")
                ws.cell(row=row_idx, column=8, value=(ledger.ledger_name if ledger else "") if is_first else "")
                
                # Column 9: Items
                item_name = item.item_name if item else ""
                if not item:
                    item_name = t.voucher_name or ""
                ws.cell(row=row_idx, column=9, value=item_name)
                
                # Column 10 to 13
                ws.cell(row=row_idx, column=10, value=(t.paid_to or "") if is_first else "")
                ws.cell(row=row_idx, column=11, value=(t.description or "") if is_first else "")
                ws.cell(row=row_idx, column=12, value=(creator.username if creator else "") if is_first else "")
                ws.cell(row=row_idx, column=13, value=(t.approved_by or "") if is_first else "")
                
                # Column 14: Amount
                item_amt = float(item.amount) if item else float(t.amount)
                ws.cell(row=row_idx, column=14, value=item_amt)
                
                # Column 15: Total
                ws.cell(row=row_idx, column=15, value=voucher_total if is_first else "")

                for col in range(1, 16):
                    c = ws.cell(row=row_idx, column=col)
                    c.border = border
                    if col in (1, 2, 4, 6, 7):
                        c.alignment = center
                    elif col in (14, 15):
                        c.alignment = right
                        c.number_format = '#,##0.00'
                    else:
                        c.alignment = left
                row_idx += 1

            if num_items > 1:
                end_row = row_idx - 1
                cols_to_merge = [1, 2, 3, 4, 5, 6, 7, 8, 10, 11, 12, 13, 15]
                for col in cols_to_merge:
                    ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
                    merged_cell = ws.cell(row=start_row, column=col)
                    if col in (1, 2, 4, 6, 7):
                        merged_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    elif col == 15:
                        merged_cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        merged_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            sno += 1

        # Total row
        total_cell = ws.cell(row=row_idx, column=13, value="Total")
        total_cell.font = Font(bold=True)
        total_cell.alignment = right
        total_cell.border = border

        grand_total = sum(
            sum(float(item.amount) for item in t.items) if getattr(t, 'items', None) else float(t.amount)
            for t in transactions
        )
        
        amt_cell = ws.cell(row=row_idx, column=14, value=grand_total)
        amt_cell.font = Font(bold=True)
        amt_cell.alignment = right
        amt_cell.number_format = '#,##0.00'
        amt_cell.border = border

        tot_cell = ws.cell(row=row_idx, column=15, value=grand_total)
        tot_cell.font = Font(bold=True)
        tot_cell.alignment = right
        tot_cell.number_format = '#,##0.00'
        tot_cell.border = border

        for col in range(1, 13):
            ws.cell(row=row_idx, column=col).border = border

        # Column widths
        widths = [6, 12, 25, 25, 25, 14, 12, 20, 20, 20, 35, 18, 18, 12, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Petty_Cash_Report_{academic_year}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logger.error(f"Error exporting petty cash excel: {str(e)}")
        return jsonify({"message": str(e)}), 500

#---------------------------
#  Month-Wise Ledger
# --------------------------
@petty_cash_report_bp.route('/month-wise-ledger', methods = ['Get'])
@token_required
def month_wise_ledger(current_user):
    try:
        from helpers import has_permission, validate_cross_branch_access
        # Feature Permission
        if not has_permission(current_user, "fees.fee.petty-cash-report", "read"):
            return jsonify({"message": "Unauthorized"}), 403

        academic_year = request.headers.get("X-Academic-Year", "2024-2025")

        branch_val = request.args.get("branch_id") or request.headers.get("X-Branch") or current_user.branch
        from routes.petty_cash_routes import resolve_branch_id
        branch_id = resolve_branch_id(branch_val)

        # Branch Scope
        if branch_id:
            is_valid, msg = validate_cross_branch_access(current_user, source_branch_id=branch_id)
            if not is_valid:
                return jsonify({"message": msg}), 403

        if not branch_id:
            return jsonify({"message": "Valid Branch required"}), 400

        fy_start_year = int(academic_year.split('-')[0])
        months = [
            ("Apr", fy_start_year, 4),
            ("May", fy_start_year, 5),
            ("Jun", fy_start_year, 6),
            ("Jul", fy_start_year, 7),
            ("Aug", fy_start_year, 8),
            ("Sep", fy_start_year, 9),
            ("Oct", fy_start_year, 10),
            ("Nov", fy_start_year, 11),
            ("Dec", fy_start_year, 12),
            ("Jan", fy_start_year + 1, 1),
            ("Feb", fy_start_year + 1, 2),
            ("Mar", fy_start_year + 1, 3),
        ]

        # Calculate Opening Balance: everything strictly before April of fy_start_year
        # Note: Usually Opening Balance is brought forward from previous years. 
        # We can sum all allocations < fy_start_year-04-01 minus all expenses < fy_start_year-04-01
        start_date = datetime(fy_start_year, 4, 1).date()
        
        prev_allocations = db.session.query(func.coalesce(func.sum(PettyCashFundAllocation.amount), 0)).filter(
            PettyCashFundAllocation.branch_id == branch_id,
            PettyCashFundAllocation.allocation_date < start_date,
            PettyCashFundAllocation.is_active == True,
            PettyCashFundAllocation.approval_status == 'Approved'
        ).scalar() or 0
        
        prev_expenses = db.session.query(func.coalesce(func.sum(PettyCash.amount), 0)).filter(
            PettyCash.branch_id == branch_id,
            PettyCash.transaction_date < start_date,
            PettyCash.voucher_type.in_(['Payment', 'Payments']),
            PettyCash.is_active == True,
            PettyCash.approval_status == 'Approved'
        ).scalar() or 0
        
        opening_balance = float(prev_allocations) - float(prev_expenses)
        
        result = []
        running_balance = opening_balance
        
        # We add the opening balance as the first row
        result.append({
            "particulars": "Opening Balance",
            "debit": 0,
            "credit": 0,
            "cash_in_hand": running_balance,
            "is_opening": True
        })
        
        # Current Year Transactions
        for label, yr, mn in months:
            # Allocations for month
            debit = db.session.query(func.coalesce(func.sum(PettyCashFundAllocation.amount), 0)).filter(
                PettyCashFundAllocation.branch_id == branch_id,
                PettyCashFundAllocation.academic_year == academic_year,
                PettyCashFundAllocation.is_active == True,
                PettyCashFundAllocation.approval_status == 'Approved',
                extract('month', PettyCashFundAllocation.allocation_date) == mn,
                extract('year', PettyCashFundAllocation.allocation_date) == yr
            ).scalar() or 0
            
            # Expenses for month
            credit = db.session.query(func.coalesce(func.sum(PettyCash.amount), 0)).filter(
                PettyCash.branch_id == branch_id,
                PettyCash.academic_year == academic_year,
                PettyCash.voucher_type.in_(['Payment', 'Payments']),
                PettyCash.is_active == True,
                PettyCash.approval_status == 'Approved',
                extract('month', PettyCash.transaction_date) == mn,
                extract('year', PettyCash.transaction_date) == yr
            ).scalar() or 0
            
            debit = float(debit)
            credit = float(credit)
            
            running_balance = running_balance + debit - credit
            
            # Use label "Apr-2026" etc
            month_label = f"{label}-{yr}"
            
            result.append({
                "particulars": month_label,
                "debit": debit,
                "credit": credit,
                "cash_in_hand": running_balance,
                "is_opening": False
            })

        return jsonify(result), 200
    except Exception as e:
        logger.error(f"Error in month wise ledger: {str(e)}")
        return jsonify({"message": str(e)}), 500


# -----------------------------
# DETAILED LEDGER
# -----------------------------
@petty_cash_report_bp.route('/ledger-details', methods=['GET'])
@token_required
def ledger_details(current_user):
    try:
        from helpers import has_permission, validate_cross_branch_access
        # Feature Permission
        if not has_permission(current_user, "fees.fee.petty-cash-report", "read"):
            return jsonify({"message": "Unauthorized"}), 403

        academic_year = request.headers.get("X-Academic-Year", "2024-2025")
        
        branch_val = request.args.get("branch_id") or request.headers.get("X-Branch") or current_user.branch
        from routes.petty_cash_routes import resolve_branch_id
        branch_id = resolve_branch_id(branch_val)

        # Branch Scope
        if branch_id:
            is_valid, msg = validate_cross_branch_access(current_user, source_branch_id=branch_id)
            if not is_valid:
                return jsonify({"message": msg}), 403

        if not branch_id:
            return jsonify({"message": "Valid Branch required"}), 400

        # Fetch Allocations
        allocations = PettyCashFundAllocation.query.filter_by(
            branch_id=branch_id, 
            academic_year=academic_year, 
            is_active=True,
            approval_status='Approved'
        ).all()
        
        # Fetch Expenses
        expenses = PettyCash.query.filter_by(
            branch_id=branch_id, 
            academic_year=academic_year, 
            is_active=True,
            approval_status='Approved'
        ).filter(PettyCash.voucher_type.in_(['Payment', 'Payments'])).all()
        
        combined = []
        
        for a in allocations:
            combined.append({
                "date_obj": a.allocation_date,
                "date": a.allocation_date.isoformat(),
                "voucher_no": f"ALLOC-{a.id}",
                "voucher_type": "Fund Allocation",
                "ledger_type": "Direct",
                "ledger_name": "Head Office Account",
                "narration": a.remark or "Fund Allocated to Branch",
                "debit": float(a.amount),
                "credit": 0.0
            })
            
        for e in expenses:
            # We map 'Received' as Debit and 'Payment' as Credit
            is_debit = e.voucher_type == 'Received'
            items_list = []
            if getattr(e, 'items', None):
                for i in e.items:
                    items_list.append({
                        "item_name": i.item_name,
                        "amount": float(i.amount)
                    })
            combined.append({
                "date_obj": e.transaction_date,
                "date": e.transaction_date.isoformat(),
                "voucher_no": e.voucher_name or f"C{e.id:05d}",
                "voucher_type": e.voucher_type,
                "ledger_type": e.ledger.ledger_type if e.ledger else "",
                "ledger_name": e.ledger.ledger_name if e.ledger else "",
                "narration": e.description or "",
                "debit": float(e.amount) if is_debit else 0.0,
                "credit": 0.0 if is_debit else float(e.amount),
                "items": items_list
            })
            
        # Sort by date
        combined.sort(key=lambda x: x["date_obj"])
        
        # Clean up date_obj and add running balance if we want, but frontend can do running balance easily too.
        # Actually frontend might want S.No which is just index + 1
        return jsonify([{k: v for k, v in row.items() if k != "date_obj"} for row in combined]), 200
        
    except Exception as e:
        logger.error(f"Error in ledger details: {str(e)}")
        return jsonify({"message": str(e)}), 500